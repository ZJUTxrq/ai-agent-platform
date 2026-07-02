from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Any, Mapping
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_REQUIREMENT_REVIEW_DOCUMENT_EXPORT_ROWS = 10000
MAX_REQUIREMENT_REVIEW_RESULT_EXPORT_ROWS = 10000
REQUIREMENT_REVIEW_EXPORT_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _coerce_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _json_dump(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, indent=2, default=str)


def _join_lines(value: Any) -> str:
    if not isinstance(value, list):
        return ""
    return "\n".join(_coerce_text(item) for item in value if _coerce_text(item))


def _apply_widths(sheet: Any, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _append_meta_sheet(
    workbook: Workbook,
    *,
    project_id: str,
    total: int,
    filters: Mapping[str, Any],
) -> None:
    sheet = workbook.create_sheet("Export Meta")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    rows = [
        ["Field", "Value"],
        ["project_id", project_id],
        ["exported_at", datetime.now().isoformat(timespec="seconds")],
        ["total", str(total)],
    ]
    for key, value in filters.items():
        rows.append([key, _coerce_text(value)])
    for row_index, values in enumerate(rows, start=1):
        sheet.append(values)
        if row_index == 1:
            for cell in sheet[row_index]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = wrap_alignment
        else:
            for cell in sheet[row_index]:
                cell.alignment = wrap_alignment
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 56


def build_requirement_review_content_disposition(filename: str) -> str:
    encoded = quote(filename)
    return f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded}"


def _document_export_filename(project_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"requirement-review-documents-{(project_id or 'project')[:8]}-{timestamp}.xlsx"


def _result_export_filename(project_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"requirement-review-results-{(project_id or 'project')[:8]}-{timestamp}.xlsx"


def build_requirement_review_documents_workbook(
    *,
    project_id: str,
    documents: list[dict[str, Any]],
    filters: Mapping[str, Any],
) -> tuple[str, bytes]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Documents"
    headers = [
        "Document ID",
        "Filename",
        "Batch ID",
        "Thread ID",
        "Content Type",
        "Source Kind",
        "Parse Status",
        "Summary",
        "Parsed Text",
        "Structured Data",
        "Provenance",
        "Error",
        "Created At",
        "Updated At",
    ]
    widths = [38, 28, 38, 28, 20, 16, 16, 40, 60, 36, 36, 28, 22, 22]
    _apply_widths(sheet, widths)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment
    for document in documents:
        row = [
            _coerce_text(document.get("id")),
            _coerce_text(document.get("filename")),
            _coerce_text(document.get("batch_id")),
            _coerce_text(document.get("thread_id")),
            _coerce_text(document.get("content_type")),
            _coerce_text(document.get("source_kind")),
            _coerce_text(document.get("parse_status")),
            _coerce_text(document.get("summary_for_model")),
            _coerce_text(document.get("parsed_text")),
            _json_dump(document.get("structured_data")),
            _json_dump(document.get("provenance")),
            _json_dump(document.get("error")),
            _coerce_text(document.get("created_at")),
            _coerce_text(document.get("updated_at")),
        ]
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment
    _append_meta_sheet(workbook, project_id=project_id, total=len(documents), filters=filters)
    output = BytesIO()
    workbook.save(output)
    return _document_export_filename(project_id), output.getvalue()


def build_requirement_review_results_workbook(
    *,
    project_id: str,
    results: list[dict[str, Any]],
    filters: Mapping[str, Any],
) -> tuple[str, bytes]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    headers = [
        "Result ID",
        "Requirement Summary",
        "Batch ID",
        "Thread ID",
        "Review Score",
        "Quality Gate",
        "Generation Policy",
        "Generation Policy Reason",
        "Document IDs",
        "Dimension Scores",
        "Key Findings",
        "Major Risks",
        "Missing Or Ambiguous Items",
        "Suggestions To Improve",
        "Assumptions",
        "Raw Result",
        "Created At",
        "Updated At",
    ]
    widths = [38, 48, 38, 28, 14, 16, 28, 48, 28, 32, 32, 32, 32, 32, 28, 36, 22, 22]
    _apply_widths(sheet, widths)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment
    for result in results:
        row = [
            _coerce_text(result.get("id")),
            _coerce_text(result.get("requirement_summary")),
            _coerce_text(result.get("batch_id")),
            _coerce_text(result.get("thread_id")),
            _coerce_text(result.get("review_score")),
            _coerce_text(result.get("quality_gate")),
            _coerce_text(result.get("generation_policy")),
            _coerce_text(result.get("generation_policy_reason")),
            _join_lines(result.get("document_ids")),
            _json_dump(result.get("dimension_scores")),
            _join_lines(result.get("key_findings")),
            _join_lines(result.get("major_risks")),
            _join_lines(result.get("missing_or_ambiguous_items")),
            _join_lines(result.get("suggestions_to_improve")),
            _join_lines(result.get("assumptions")),
            _json_dump(result.get("raw_result")),
            _coerce_text(result.get("created_at")),
            _coerce_text(result.get("updated_at")),
        ]
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment
    _append_meta_sheet(workbook, project_id=project_id, total=len(results), filters=filters)
    output = BytesIO()
    workbook.save(output)
    return _result_export_filename(project_id), output.getvalue()
